# Databricks notebook source
# DBTITLE 1,NSO Step 2 - Parse Excel Workbooks Only
# MAGIC %md
# MAGIC # Step 2: Parse NSO Excel workbooks only
# MAGIC
# MAGIC This notebook intentionally ignores HTML/DOC/DOCX narrative files. The NSO statistical attachments are Excel workbooks, and samples in the `download_docs` volume show two important formats:
# MAGIC
# MAGIC - modern `.xlsx` ZIP/XML workbooks, usually Unicode Vietnamese, many sheets, multi-row headers, merged header cells, sparse trailing columns
# MAGIC - legacy `.xls` BIFF workbooks from older years, often stored with `.xls` extension but `attachment_type = 'xlsx'` from the crawler; these require `xlrd` and may contain legacy Vietnamese font encodings
# MAGIC
# MAGIC Output remains compatible with downstream notebooks via `market_data.nso.parsed_workbooks_raw.raw_cells_json`.

# COMMAND ----------

import importlib.util
import json
import re
import subprocess
import sys
import unicodedata
import zipfile
from datetime import datetime, date
from decimal import Decimal

from pyspark.sql.types import *

CATALOG = "market_data"
SCHEMA = "nso"

# Existing Step 3 reads this table, so keep the table name stable.
PARSED_TABLE = f"{CATALOG}.{SCHEMA}.parsed_workbooks_raw"
LOG_TABLE = f"{CATALOG}.{SCHEMA}.document_processing_log"
ATTACHMENTS_TABLE = f"{CATALOG}.{SCHEMA}.nso_report_attachments"

# Safe default: Step 2 should process Excel only and should not mutate doc/html parse status.
MARK_NON_EXCEL_SKIPPED = False
MAX_CELL_TEXT_LENGTH = 20000

spark.sql(f"CREATE SCHEMA IF NOT EXISTS {CATALOG}.{SCHEMA}")

# COMMAND ----------

# DBTITLE 1,Runtime dependencies
# Databricks runtimes often include openpyxl but not xlrd. Install missing packages in-place so
# legacy .xls files can be parsed in the same workflow task.
def ensure_package(import_name, pip_name):
    if importlib.util.find_spec(import_name) is None:
        print(f"Installing missing package: {pip_name}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", pip_name])

ensure_package("openpyxl", "openpyxl")
ensure_package("xlrd", "xlrd==2.0.1")

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter

# Some legacy NSO .xls files carry invalid UTF-16 in workbook metadata (SUPBOOK external-reference
# names), which makes xlrd's strict decode raise before any cell is read. Fall back to
# errors='replace' only when strict decoding fails; cell text in the affected files decodes clean.
def _lenient_unicode(b, enc):
    try:
        return b.decode(enc)
    except UnicodeDecodeError:
        return b.decode(enc, "replace")

for _mod in (xlrd.biffh, xlrd.book, xlrd.sheet, xlrd.formula, xlrd.timemachine):
    if hasattr(_mod, "unicode"):
        _mod.unicode = _lenient_unicode

# COMMAND ----------

# DBTITLE 1,Tables
spark.sql(f"""
CREATE TABLE IF NOT EXISTS {PARSED_TABLE} (
  attachment_id STRING NOT NULL,
  report_id STRING NOT NULL,
  report_url STRING,
  attachment_url STRING,
  filename STRING,
  sheet_index INT,
  sheet_name_raw STRING,
  sheet_name_normalized STRING,
  max_row INT,
  max_column INT,
  raw_cells_json STRING,
  parsed_timestamp TIMESTAMP
) USING DELTA
""")

# Add Excel-specific metadata without breaking older downstream readers.
for ddl in [
    "excel_file_format STRING",
    "parser_engine STRING",
    "parser_version STRING",
    "non_empty_cell_count INT",
    "first_non_empty_row INT",
    "first_non_empty_column INT",
    "merged_ranges_json STRING",
    "parse_warnings_json STRING",
    "detected_text_encoding STRING",
]:
    try:
        spark.sql(f"ALTER TABLE {PARSED_TABLE} ADD COLUMNS ({ddl})")
    except Exception as e:
        if "already exists" not in str(e).lower() and "duplicate" not in str(e).lower():
            print(f"Column add skipped/failed for {ddl}: {e}")

# COMMAND ----------

# DBTITLE 1,Helpers
def norm_text(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        # Keep numeric values stable but avoid ugly integer-looking floats.
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, Decimal):
        return format(value, "f")
    text = str(value)
    text = text.replace(" ", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None

_TCVN3_CHAR_MAP = {
    "µ":"à", "¸":"á", "¶":"ả", "·":"ã", "¹":"ạ",
    "¨":"ă", "¾":"ắ", "»":"ằ", "¼":"ẳ", "½":"ẵ", "Æ":"ặ",
    "©":"â", "Ê":"ấ", "Ç":"ầ", "È":"ẩ", "É":"ẫ", "Ë":"ậ",
    "Ì":"è", "Ð":"é", "Î":"ẻ", "Ï":"ẽ", "Ñ":"ẹ",
    "ª":"ê", "Õ":"ế", "Ò":"ề", "Ó":"ể", "Ô":"ễ", "Ö":"ệ",
    "×":"ì", "Ý":"í", "Ø":"ỉ", "Ü":"ĩ", "Þ":"ị",
    "ß":"ò", "ã":"ó", "á":"ỏ", "â":"õ", "ä":"ọ",
    "«":"ô", "è":"ố", "å":"ồ", "æ":"ổ", "ç":"ỗ", "é":"ộ",
    "¬":"ơ", "í":"ớ", "ê":"ờ", "ë":"ở", "ì":"ỡ", "î":"ợ",
    "ï":"ù", "ó":"ú", "ñ":"ủ", "ò":"ũ", "ô":"ụ",
    "­":"ư", "ø":"ứ", "õ":"ừ", "ö":"ử", "÷":"ữ", "ù":"ự",
    "ú":"ỳ", "ý":"ỷ", "û":"ỹ", "þ":"ỵ", "§":"đ",
    "¡":"Ă", "¢":"Â", "£":"Ê", "¤":"Ô", "¥":"Ơ", "¦":"Ư", "®":"Đ",
}
TCVN3_TO_UNICODE = str.maketrans(_TCVN3_CHAR_MAP)

# Unambiguous evidence characters: pure symbols/punctuation in any normal font, never valid
# standalone Vietnamese or Latin letters. The rest of _TCVN3_CHAR_MAP's keys (the second
# tone-group half -- accented a/e/i/o/u letters like U+00EA "e-circumflex", U+00ED "i-acute")
# double as legitimate accented Latin-1 letters -- e.g. U+00ED is both a real standalone
# Vietnamese vowel AND a TCVN3-glyph key for a different letter -- so they can't be used as
# detection evidence without false positives. They are still translated once a sheet is
# classified TCVN3 from the unambiguous evidence below.
TCVN3_UNAMBIGUOUS_GLYPHS = frozenset(
    "µ¸¶·¹¨¾»¼½Æ"
    "©ª«¬­®§¡¢£¤¥¦×÷"
)

def count_unambiguous_tcvn3_hits(text):
    return sum(1 for ch in text if ch in TCVN3_UNAMBIGUOUS_GLYPHS)

def detect_sheet_encoding(text_values, sample_size=300, convert_threshold=3, review_threshold=1):
    # Classify the whole sheet once from a large text sample rather than gating each cell
    # individually on a fixed keyword list. A larger sample means rare words/short cells
    # don't matter -- only a few unambiguous glyphs need to show up anywhere in the sheet.
    # A single incidental hit (e.g. a real section-reference symbol or footnote superscript)
    # is deliberately NOT auto-converted -- it's flagged for review instead, since one symbol
    # is weak evidence but three or more scattered through a sheet essentially never happens
    # in legitimate report text.
    sample = " ".join(v for v in text_values[:sample_size] if isinstance(v, str))
    hits = count_unambiguous_tcvn3_hits(sample)
    if hits >= convert_threshold:
        return "tcvn3"
    if hits >= review_threshold:
        return "possible_legacy_low_confidence"
    return "unicode"

# TCVN3 is an 8-bit encoding: genuinely TCVN3-encoded text can only decode into Latin-1-range
# characters, never into the Vietnamese letters that live above U+00FF (ă đ ĩ ơ ư and the whole
# U+1EA0-U+1EF9 tone block). A cell containing any of those is therefore already proper Unicode,
# and translating it would corrupt it (e.g. real "Khách sạn" -> "Khỏch sạn", because "á" doubles
# as the TCVN3 byte for "ỏ"). Sheets can mix both kinds of cells, so the sheet-level decision
# alone is not enough -- each cell still needs this guard.
VIETNAMESE_BEYOND_LATIN1 = frozenset("ĂăĐđĨĩŨũƠơƯư") | frozenset(
    chr(cp) for cp in range(0x1EA0, 0x1EFA)
)

def maybe_tcvn3_to_unicode(text, force=False):
    if not text:
        return text
    # `force` comes from the sheet-level detection decision (detect_sheet_encoding).
    if force:
        if any(ch in VIETNAMESE_BEYOND_LATIN1 for ch in text):
            return text
        return text.translate(TCVN3_TO_UNICODE)
    return text

def repair_vietnamese_unicode(text, force_tcvn3=False):
    if text is None:
        return None
    repaired = maybe_tcvn3_to_unicode(str(text), force=force_tcvn3)
    repaired = re.sub(r"\s+", " ", repaired.replace(" ", " ")).strip()
    return repaired or None

# Backfill corrections for normalized text from legacy rows.
VIETNAMESE_ASCII_CORRECTIONS = [
    (r"\bHang hoo khoc\b", "Hang hoa khac"),
    (r"\bhang hoo khoc\b", "hang hoa khac"),
    (r"\bHang hua\b", "Hang hoa"),
    (r"\bhang hua\b", "hang hoa"),
    (r"\bhua\b", "hoa"),
    (r"\bHua\b", "Hoa"),
    (r"\bhoo\b", "hoa"),
    (r"\bHoo\b", "Hoa"),
    (r"\bkhoc\b", "khac"),
    (r"\bKhoc\b", "Khac"),
    (r"\bcho thuo\b", "cho thue"),
    (r"\bCho thuo\b", "Cho thue"),
    (r"\bmay muc\b", "may moc"),
    (r"\bMay muc\b", "May moc"),
    (r"\bkho boi\b", "kho bai"),
    (r"\bKho boi\b", "Kho bai"),
    (r"\bcu lion quan\b", "co lien quan"),
    (r"\bCu lion quan\b", "Co lien quan"),
    (r"\bcu lien quan\b", "co lien quan"),
    (r"\bCu lien quan\b", "Co lien quan"),
    (r"\blion quan\b", "lien quan"),
    (r"\bLion quan\b", "Lien quan"),
    (r"\bcoc loai\b", "cac loai"),
    (r"\bCoc loai\b", "Cac loai"),
    (r"\bcoc dich vu\b", "cac dich vu"),
    (r"\bCoc dich vu\b", "Cac dich vu"),
    (r"\bcoc san pham\b", "cac san pham"),
    (r"\bCoc san pham\b", "Cac san pham"),
    (r"\bcoc thiet bi\b", "cac thiet bi"),
    (r"\bCoc thiet bi\b", "Cac thiet bi"),
    (r"\bBon buun\b", "Ban buon"),
    (r"\bbon buun\b", "ban buon"),
    (r"\bbon le\b", "ban le"),
    (r"\bBon le\b", "Ban le"),
    (r"\bu tu\b", "o to"),
    (r"\bU tu\b", "O to"),
    (r"\bmu tu\b", "mo to"),
    (r"\bMu tu\b", "Mo to"),
    (r"\bcu dong\b", "co dong"),
    (r"\bCu dong\b", "Co dong"),
    (r"\bNung, lom nghiep\b", "Nong, lam nghiep"),
    (r"\bnung, lom nghiep\b", "nong, lam nghiep"),
    (r"\bNung nghiep\b", "Nong nghiep"),
    (r"\bnung nghiep\b", "nong nghiep"),
    (r"\bLom nghiep\b", "Lam nghiep"),
    (r"\blom nghiep\b", "lam nghiep"),
    (r"\bphEm\b", "pham"),
    (r"\bthop\b", "thep"),
    (r"\btiou\b", "tieu"),
    (r"\bNghon\b", "Nghin"),
    (r"\bnghon\b", "nghin"),
    (r"\btonh\b", "tinh"),
    (r"\bTonh\b", "Tinh"),
    (r"\btoch\b", "tich"),
    (r"\bToch\b", "Tich"),
    (r"\bmoy\b", "may"),
    (r"\bMoy\b", "May"),
    (r"\bTri gio\b", "Tri gia"),
    (r"\btri gio\b", "tri gia"),
    (r"\bTriOu\b", "Trieu"),
    (r"\bTRiOu\b", "Trieu"),
    (r"\btriOu\b", "trieu"),
    (r"\bGioo duc va dao tao\b", "Giao duc va dao tao"),
    (r"\bgiai tro\b", "giai tri"),
    (r"\bBon buun; bon le; sua chua u tu, xe may\b", "Ban buon; ban le; sua chua o to, xe may"),
    (r"\bDich vu viec lam; du lich; cho thuo may muc thiet bi, do dung va coc dich vu ho tro khoc\b", "Dich vu viec lam; du lich; cho thue may moc thiet bi, do dung va cac dich vu ho tro khac"),
    (r"\bHua chat\b", "Hoa chat"),
    (r"\bKho dot hua long\b", "Kho dot hoa long"),
    (r"\bThanh Hua\b", "Thanh Hoa"),
    (r"\bKim loai thuong khoc\b", "Kim loai thuong khac"),
    (r"\bGiay coc loai\b", "Giay cac loai"),
    (r"\bCung cap nuoc; hoat dong quan ly va xu ly roc thai, nuoc thai\b", "Cung cap nuoc; hoat dong quan ly va xu ly rac thai, nuoc thai"),
    (r"\bSan xuat da va coc san pham cu lion quan\b", "San xuat da va cac san pham co lien quan"),
    (r"\bSan xuat san pham thuoc lo\b", "San xuat san pham thuoc la"),
    (r"\bSan xuat thuoc, hoo duoc va duoc lieu\b", "San xuat thuoc, hoa duoc va duoc lieu"),
    (r"\bHoat dong thu gom, xu ly va tieu huy roc thai; toi che phe lieu\b", "Hoat dong thu gom, xu ly va tieu huy rac thai; tai che phe lieu"),
    (r"\bBonh Duong\b", "Binh Duong"),
    (r"\bThoi Bonh\b", "Thai Binh"),
    (r"\bBonh Dinh\b", "Binh Dinh"),
    (r"\bGioo\b", "Giao"),
    (r"\bgioo\b", "giao"),
    (r"\bthuoc lo\b", "thuoc la"),
    (r"\bThuoc lo\b", "Thuoc la"),
    (r"\broc thai\b", "rac thai"),
    (r"\bRoc thai\b", "Rac thai"),
    (r"\btoi che\b", "tai che"),
    (r"\bcoc dich vu lion quan\b", "cac dich vu lien quan"),
    (r"\bDich vu kho boi va cac dich vu lien quan den ho tro van tai\b", "Dich vu kho bai va cac dich vu lien quan den ho tro van tai"),
    (r"\bBon buun va bon le; sua chua u tu, mu tu, xe may va xe cu dong co khoc\b", "Ban buon va ban le; sua chua o to, mo to, xe may va xe co dong co khac"),
    (r"\bCung cap nuoc, hoat dong quan ly va xu ly rac thai, nuoc thai\b", "Cung cap nuoc; hoat dong quan ly va xu ly rac thai, nuoc thai"),
    (r"\bNuoc tu nhion khai thoc\b", "Nuoc tu nhien khai thac"),
    (r"\bLom nghiep va dich vu co lien quan\b", "Lam nghiep va dich vu co lien quan"),
    (r"\bNung nghiep va dich vu co lien quan\b", "Nong nghiep va dich vu co lien quan"),
    (r"\bDet, trang phuc, da va coc san pham co lien quan\b", "Det, trang phuc, da va cac san pham co lien quan"),
    (r"\bDa va coc san pham co lien quan\b", "Da va cac san pham co lien quan"),
    (r"\bHoat dog thu gom\b", "Hoat dong thu gom"),
    (r"\bVan hoo\b", "Van hoa"),

    (r"\bBonh keo\b", "Banh keo"),
    (r"\bBonh quon\b", "Banh quan"),
    (r"\bbonh\b", "banh"),
    (r"\bBonh\b", "Banh"),
    (r"\bCung nghiep\b", "Cong nghiep"),
    (r"\bcung nghiep\b", "cong nghiep"),
    (r"\bcung nghe\b", "cong nghe"),
    (r"\bCung nghe\b", "Cong nghe"),
    (r"\bchuyon mun\b", "chuyen mon"),
    (r"\bChuyon mun\b", "Chuyen mon"),
    (r"\bchuyon\b", "chuyen"),
    (r"\bChuyon\b", "Chuyen"),
    (r"\bhang khung\b", "hang khong"),
    (r"\bHang khung\b", "Hang khong"),
    (r"\bkhung kho\b", "khong khi"),
    (r"\bKhung kho\b", "Khong khi"),
    (r"\bnuoc nung\b", "nuoc nong"),
    (r"\bNuoc nung\b", "Nuoc nong"),
    (r"\bChi so gio\b", "Chi so gia"),
    (r"\bchi so gio\b", "chi so gia"),
    (r"\bhanh chonh\b", "hanh chinh"),
    (r"\bHanh chonh\b", "Hanh chinh"),
    (r"\bgia donh\b", "gia dinh"),
    (r"\bGia donh\b", "Gia dinh"),
    (r"\bcoc cung viec\b", "cac cong viec"),
    (r"\bCoc cung viec\b", "Cac cong viec"),
]

def fix_vietnamese_ascii_artifacts(value):
    if value is None:
        return None
    text = str(value)
    for pattern, replacement in VIETNAMESE_ASCII_CORRECTIONS:
        text = re.sub(pattern, replacement, text)
    text = re.sub(r"\s+", " ", text.replace(" ", " ")).strip()
    return text or None


def remove_vietnamese_marks(text, force_tcvn3=False):
    if text is None:
        return None
    text = maybe_tcvn3_to_unicode(str(text), force=force_tcvn3)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    text = re.sub(r"\s+", " ", text.replace(" ", " ")).strip()
    return fix_vietnamese_ascii_artifacts(text)

def normalize_name(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def is_excel_attachment(row_dict):
    filename = (row_dict.get("filename") or row_dict.get("attachment_url") or "").lower()
    typ = (row_dict.get("attachment_type") or "").lower()
    return typ in {"xls", "xlsx", "excel", "spreadsheet"} or filename.endswith((".xls", ".xlsx", ".xlsm"))

def excel_format_from_path(path, filename=None):
    name = (filename or path or "").lower()
    try:
        if zipfile.is_zipfile(path):
            return "xlsx"
        with open(path, "rb") as f:
            magic = f.read(8)
        if magic.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
            return "xls"
    except Exception:
        pass
    if name.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    return "unknown"

def cell_ref(row_index, column_index):
    return f"{get_column_letter(column_index)}{row_index}"

def trim_cell_text(text):
    if text is None:
        return None
    if len(text) > MAX_CELL_TEXT_LENGTH:
        return text[:MAX_CELL_TEXT_LENGTH] + "…[truncated]"
    return text

# COMMAND ----------

# DBTITLE 1,Parsers
def finalize_sheet_cells(raw_cells, text_values):
    # Decide the sheet's text encoding once from the full cell-text sample, then apply that
    # single decision uniformly to every cell -- instead of gating each cell individually on
    # a fixed keyword list (see detect_sheet_encoding / TCVN3_UNAMBIGUOUS_GLYPHS above).
    sheet_encoding = detect_sheet_encoding(text_values)
    force_tcvn3 = sheet_encoding == "tcvn3"
    warnings = []
    if force_tcvn3:
        warnings.append("legacy_tcvn3_font_detected_and_converted")
    elif sheet_encoding == "possible_legacy_low_confidence":
        # One or two TCVN3 glyphs is too weak for automatic conversion, but still
        # worth surfacing in the parse warnings.
        warnings.append("possible_legacy_font_low_confidence_not_converted")
    cells = []
    for rc in raw_cells:
        v = rc["cell_value_raw"]
        cells.append({
            **rc,
            "cell_value_unicode": repair_vietnamese_unicode(v, force_tcvn3=force_tcvn3),
            "cell_value_normalized": remove_vietnamese_marks(v, force_tcvn3=force_tcvn3),
        })
    return cells, sheet_encoding, warnings

def _cell_indent(cell):
    """openpyxl indent level as an int, defensively -- never fail a parse over styling."""
    try:
        return int(getattr(cell.alignment, "indent", 0) or 0)
    except Exception:
        return 0

def _xls_indent(book, sheet, rx, cx):
    """xlrd indent level, or 0 when formatting_info could not be loaded for this workbook."""
    try:
        xf = book.xf_list[sheet.cell_xf_index(rx, cx)]
        return int(getattr(xf.alignment, "indent_level", 0) or 0)
    except Exception:
        return 0

def parse_xlsx_workbook(path):
    workbook_rows = []
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            raw_cells = []
            text_values = []
            max_row = 0
            max_col = 0
            first_row = None
            first_col = None
            for row in ws.iter_rows():
                for cell in row:
                    value = norm_text(cell.value)
                    if value is None:
                        continue
                    value = trim_cell_text(value)
                    r = int(cell.row)
                    c = int(cell.column)
                    max_row = max(max_row, r)
                    max_col = max(max_col, c)
                    first_row = r if first_row is None else min(first_row, r)
                    first_col = c if first_col is None else min(first_col, c)
                    text_values.append(value)
                    raw_cells.append({
                        "row_index": r,
                        "column_index": c,
                        "cell_ref": cell.coordinate,
                        "cell_value_raw": value,
                        # Excel indent level is the structural parent/child signal in NSO sheets:
                        # a province or product sits at indent 0 and its ownership-sector or
                        # transport-mode breakdown at indent 1. It is a style attribute, not
                        # leading spaces, so it is lost unless captured here. Step 3 uses it to
                        # keep repeated child labels distinguishable.
                        "indent_level": _cell_indent(cell),
                    })
            cells, sheet_encoding, warnings = finalize_sheet_cells(raw_cells, text_values)
            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            workbook_rows.append({
                "sheet_index": sheet_index,
                "sheet_name_raw": ws.title,
                "sheet_name_normalized": normalize_name(ws.title),
                "max_row": max_row or ws.max_row,
                "max_column": max_col or ws.max_column,
                "raw_cells_json": json.dumps(cells, ensure_ascii=False),
                "excel_file_format": "xlsx",
                "parser_engine": "openpyxl",
                "parser_version": getattr(openpyxl, "__version__", None),
                "non_empty_cell_count": len(cells),
                "first_non_empty_row": first_row,
                "first_non_empty_column": first_col,
                "merged_ranges_json": json.dumps(merged_ranges, ensure_ascii=False),
                "parse_warnings_json": json.dumps(warnings, ensure_ascii=False),
                "detected_text_encoding": sheet_encoding,
            })
    finally:
        wb.close()
    return workbook_rows

def parse_xls_workbook(path):
    workbook_rows = []
    # formatting_info=True is what exposes XF records, and therefore indent levels — the
    # parent/child signal Step 3 needs. It is heavier and unsupported for a few workbooks, so
    # fall back to the previous behaviour rather than failing the parse; indent then reads 0.
    try:
        book = xlrd.open_workbook(path, on_demand=False, formatting_info=True)
    except Exception:
        book = xlrd.open_workbook(path, on_demand=True, formatting_info=False)
    try:
        datemode = book.datemode
        for sheet_index, sh in enumerate(book.sheets(), start=1):
            raw_cells = []
            text_values = []
            max_row = 0
            max_col = 0
            first_row = None
            first_col = None
            date_warnings = []
            for rx in range(sh.nrows):
                for cx in range(sh.ncols):
                    cell = sh.cell(rx, cx)
                    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        continue
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, datemode)
                        except Exception:
                            date_warnings.append(f"date_decode_failed:{rx+1}:{cx+1}")
                    text = norm_text(value)
                    if text is None:
                        continue
                    text = trim_cell_text(text)
                    r = rx + 1
                    c = cx + 1
                    max_row = max(max_row, r)
                    max_col = max(max_col, c)
                    first_row = r if first_row is None else min(first_row, r)
                    first_col = c if first_col is None else min(first_col, c)
                    text_values.append(text)
                    raw_cells.append({
                        "row_index": r,
                        "column_index": c,
                        "cell_ref": cell_ref(r, c),
                        "cell_value_raw": text,
                        "indent_level": _xls_indent(book, sh, rx, cx),
                    })
            cells, sheet_encoding, warnings = finalize_sheet_cells(raw_cells, text_values)
            warnings = date_warnings + warnings
            workbook_rows.append({
                "sheet_index": sheet_index,
                "sheet_name_raw": sh.name,
                "sheet_name_normalized": normalize_name(sh.name),
                "max_row": max_row or sh.nrows,
                "max_column": max_col or sh.ncols,
                "raw_cells_json": json.dumps(cells, ensure_ascii=False),
                "excel_file_format": "xls",
                "parser_engine": "xlrd",
                "parser_version": getattr(xlrd, "__version__", None),
                "non_empty_cell_count": len(cells),
                "first_non_empty_row": first_row,
                "first_non_empty_column": first_col,
                "merged_ranges_json": json.dumps([], ensure_ascii=False),
                "parse_warnings_json": json.dumps(warnings, ensure_ascii=False),
                "detected_text_encoding": sheet_encoding,
            })
    finally:
        book.release_resources()
    return workbook_rows

def parse_excel_workbook(path, filename=None):
    fmt = excel_format_from_path(path, filename)
    if fmt == "xlsx":
        return parse_xlsx_workbook(path)
    if fmt == "xls":
        return parse_xls_workbook(path)
    raise ValueError(f"Unsupported or unrecognized Excel format for {filename or path}")

# COMMAND ----------

# DBTITLE 1,Select pending Excel attachments
pending = spark.sql(f"""
SELECT l.*, a.local_path AS attachment_local_path
FROM {LOG_TABLE} l
LEFT JOIN {ATTACHMENTS_TABLE} a ON l.attachment_id = a.attachment_id
WHERE l.download_status = 'success'
  AND (l.parse_status IS NULL OR l.parse_status IN ('pending','failed'))
  AND (
    lower(coalesce(l.attachment_type, '')) IN ('xls','xlsx','excel','spreadsheet')
    OR lower(coalesce(l.filename, '')) RLIKE '[.](xls|xlsx|xlsm)$'
    OR lower(coalesce(l.attachment_url, '')) RLIKE '[.](xls|xlsx|xlsm)($|[?])'
  )
""").collect()

print(f"Pending Excel attachments: {len(pending)}")

# COMMAND ----------

# DBTITLE 1,Parse workbooks
workbook_rows = []
status_rows = []

for i, row in enumerate(pending, start=1):
    d = row.asDict()
    local_path = d.get("attachment_local_path") or d.get("local_path")
    if i == 1 or i % 25 == 0:
        print(f"Parsing Excel attachment {i}/{len(pending)}: {d.get('filename')} -> {local_path}")
    try:
        if not local_path:
            raise ValueError("Missing local_path for downloaded Excel attachment")
        sheets = parse_excel_workbook(local_path, d.get("filename"))
        parsed_at = datetime.utcnow()
        for sheet in sheets:
            workbook_rows.append({
                "attachment_id": d["attachment_id"],
                "report_id": d["report_id"],
                "report_url": d.get("report_url"),
                "attachment_url": d.get("attachment_url"),
                "filename": d.get("filename"),
                "parsed_timestamp": parsed_at,
                **sheet,
            })
        status_rows.append({
            "attachment_id": d["attachment_id"],
            "parse_status": "success",
            "parse_timestamp": parsed_at,
            "parse_error_message": None,
            "extraction_status": "pending",
            "updated_at": parsed_at,
        })
    except Exception as e:
        failed_at = datetime.utcnow()
        status_rows.append({
            "attachment_id": d["attachment_id"],
            "parse_status": "failed",
            "parse_timestamp": failed_at,
            "parse_error_message": str(e)[:1000],
            "extraction_status": None,
            "updated_at": failed_at,
        })

print(f"Parsed sheet rows prepared: {len(workbook_rows)}")
print(f"Status rows prepared: {len(status_rows)}")

# COMMAND ----------

# DBTITLE 1,Write parsed sheets and statuses
wb_schema = StructType([
    StructField("attachment_id", StringType(), False),
    StructField("report_id", StringType(), False),
    StructField("report_url", StringType(), True),
    StructField("attachment_url", StringType(), True),
    StructField("filename", StringType(), True),
    StructField("sheet_index", IntegerType(), True),
    StructField("sheet_name_raw", StringType(), True),
    StructField("sheet_name_normalized", StringType(), True),
    StructField("max_row", IntegerType(), True),
    StructField("max_column", IntegerType(), True),
    StructField("raw_cells_json", StringType(), True),
    StructField("parsed_timestamp", TimestampType(), True),
    StructField("excel_file_format", StringType(), True),
    StructField("parser_engine", StringType(), True),
    StructField("parser_version", StringType(), True),
    StructField("non_empty_cell_count", IntegerType(), True),
    StructField("first_non_empty_row", IntegerType(), True),
    StructField("first_non_empty_column", IntegerType(), True),
    StructField("merged_ranges_json", StringType(), True),
    StructField("parse_warnings_json", StringType(), True),
    StructField("detected_text_encoding", StringType(), True),
])

status_schema = StructType([
    StructField("attachment_id", StringType(), False),
    StructField("parse_status", StringType(), True),
    StructField("parse_timestamp", TimestampType(), True),
    StructField("parse_error_message", StringType(), True),
    StructField("extraction_status", StringType(), True),
    StructField("updated_at", TimestampType(), True),
])

if workbook_rows:
    spark.createDataFrame(workbook_rows, wb_schema).createOrReplaceTempView("new_parsed_workbooks")
    spark.sql(f"""
    MERGE INTO {PARSED_TABLE} AS target
    USING new_parsed_workbooks AS source
    ON target.attachment_id = source.attachment_id AND target.sheet_index = source.sheet_index
    WHEN MATCHED THEN UPDATE SET *
    WHEN NOT MATCHED THEN INSERT *
    """)

if status_rows:
    spark.createDataFrame(status_rows, status_schema).createOrReplaceTempView("parse_status_updates")
    spark.sql(f"""
    MERGE INTO {LOG_TABLE} AS target
    USING parse_status_updates AS source
    ON target.attachment_id = source.attachment_id
    WHEN MATCHED THEN UPDATE SET
      parse_status = source.parse_status,
      parse_timestamp = source.parse_timestamp,
      parse_error_message = source.parse_error_message,
      extraction_status = CASE
        WHEN source.parse_status = 'success' THEN source.extraction_status
        ELSE target.extraction_status
      END,
      updated_at = source.updated_at
    """)

# COMMAND ----------

# DBTITLE 1,Optional non-Excel status handling
if MARK_NON_EXCEL_SKIPPED:
    spark.sql(f"""
    UPDATE {LOG_TABLE}
    SET parse_status = 'skipped_non_excel',
        parse_timestamp = current_timestamp(),
        parse_error_message = 'Step 2 is Excel-only; non-Excel narrative files intentionally skipped',
        updated_at = current_timestamp()
    WHERE download_status = 'success'
      AND (parse_status IS NULL OR parse_status IN ('pending','failed'))
      AND NOT (
        lower(coalesce(attachment_type, '')) IN ('xls','xlsx','excel','spreadsheet')
        OR lower(coalesce(filename, '')) RLIKE '[.](xls|xlsx|xlsm)$'
        OR lower(coalesce(attachment_url, '')) RLIKE '[.](xls|xlsx|xlsm)($|[?])'
      )
    """)

# COMMAND ----------

# DBTITLE 1,QC summaries
display(spark.sql(f"""
SELECT
  attachment_type,
  parse_status,
  COUNT(*) AS attachments
FROM {LOG_TABLE}
GROUP BY attachment_type, parse_status
ORDER BY attachment_type, parse_status
"""))

display(spark.sql(f"""
SELECT
  excel_file_format,
  parser_engine,
  COUNT(DISTINCT attachment_id) AS workbooks,
  COUNT(*) AS sheets,
  SUM(non_empty_cell_count) AS non_empty_cells
FROM {PARSED_TABLE}
GROUP BY excel_file_format, parser_engine
ORDER BY excel_file_format, parser_engine
"""))

# Sheet-level TCVN3 detection outcomes. `possible_legacy_low_confidence` sheets contain one or
# two unambiguous legacy-glyph symbols -- too weak to safely auto-convert (could be a real
# footnote/section-reference symbol) -- and should be reviewed manually rather than papered
# over with another ad-hoc word-list patch.
display(spark.sql(f"""
SELECT
  detected_text_encoding,
  COUNT(DISTINCT attachment_id) AS workbooks,
  COUNT(*) AS sheets
FROM {PARSED_TABLE}
GROUP BY detected_text_encoding
ORDER BY detected_text_encoding
"""))

display(spark.sql(f"""
SELECT
  attachment_id,
  filename,
  sheet_index,
  sheet_name_raw,
  excel_file_format,
  detected_text_encoding,
  max_row,
  max_column,
  non_empty_cell_count,
  parse_warnings_json
FROM {PARSED_TABLE}
WHERE detected_text_encoding = 'possible_legacy_low_confidence'
ORDER BY parsed_timestamp DESC, attachment_id, sheet_index
LIMIT 50
"""))

display(spark.sql(f"""
SELECT
  attachment_id,
  filename,
  sheet_index,
  sheet_name_raw,
  excel_file_format,
  max_row,
  max_column,
  non_empty_cell_count,
  parse_warnings_json
FROM {PARSED_TABLE}
ORDER BY parsed_timestamp DESC, attachment_id, sheet_index
LIMIT 50
"""))
